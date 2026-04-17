"""Excel exporter — generates .xlsx tracker files using openpyxl."""

import json
import logging
from datetime import datetime
from pathlib import Path
from sqlalchemy.orm import Session
from models.application import Application
from models.job import Job
from config import settings

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export job tracking data to Excel (.xlsx)."""

    def export(self, db: Session) -> str:
        """Export all applications to an Excel file."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Job Tracker"

            # Headers
            headers = [
                "Company", "Role", "Location", "Source", "Job URL",
                "Resume Link", "ATS Score", "Match Score", "Status",
                "Date Added", "Applied Date", "Interview Date",
                "Missing Skills", "Notes"
            ]

            # Header styling
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Data rows
            applications = (
                db.query(Application)
                .join(Job)
                .order_by(Application.date_added.desc())
                .all()
            )

            # Score color fills
            green_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
            yellow_fill = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
            red_fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")

            for row_idx, app in enumerate(applications, 2):
                job = app.job

                ws.cell(row=row_idx, column=1, value=job.company if job else "")
                ws.cell(row=row_idx, column=2, value=job.title if job else "")
                ws.cell(row=row_idx, column=3, value=job.location if job else "")
                ws.cell(row=row_idx, column=4, value=job.source if job else "")
                ws.cell(row=row_idx, column=5, value=job.job_url if job else "")
                ws.cell(row=row_idx, column=6, value=app.resume_pdf_path or "")

                # ATS Score with color
                score_cell = ws.cell(row=row_idx, column=7, value=app.ats_score or 0)
                if app.ats_score:
                    if app.ats_score >= 80:
                        score_cell.fill = green_fill
                    elif app.ats_score >= 60:
                        score_cell.fill = yellow_fill
                    else:
                        score_cell.fill = red_fill

                # Match score as percentage
                match = job.match_score if job and job.match_score else 0
                ws.cell(row=row_idx, column=8, value=f"{match*100:.0f}%")

                ws.cell(row=row_idx, column=9, value=app.status or "not_applied")
                ws.cell(row=row_idx, column=10, value=str(app.date_added) if app.date_added else "")
                ws.cell(row=row_idx, column=11, value=str(app.applied_date) if app.applied_date else "")
                ws.cell(row=row_idx, column=12, value=str(app.interview_date) if app.interview_date else "")

                missing = json.loads(app.missing_skills) if app.missing_skills else []
                ws.cell(row=row_idx, column=13, value=", ".join(missing))
                ws.cell(row=row_idx, column=14, value=app.notes or "")

                # Apply borders
                for col in range(1, 15):
                    ws.cell(row=row_idx, column=col).border = thin_border

            # Auto-fit column widths (approximate)
            col_widths = [18, 22, 16, 10, 40, 30, 10, 10, 14, 18, 14, 18, 25, 30]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + i) if i <= 26 else ""].width = width

            # Freeze header row
            ws.freeze_panes = "A2"

            # Save
            today = datetime.now().strftime("%Y-%m-%d")
            export_dir = settings.data_path / "exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            file_path = export_dir / f"job_tracker_{today}.xlsx"
            wb.save(str(file_path))

            logger.info(f"Excel export saved: {file_path}")
            return str(file_path)

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise


excel_exporter = ExcelExporter()
